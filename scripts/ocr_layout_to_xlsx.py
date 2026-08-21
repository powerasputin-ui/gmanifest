"""Reconstructs an approximate visual layout from PaddleOCR's per-line
bounding boxes into a real, editable .xlsx — the spreadsheet analogue of
ocr_layout_to_docx.py. Same evidence, same "PDF -> same file back, no LLM
drawing anything" principle, just a different render target: a general copy
of the source document, not a fixed-field form (that's export-sgt, a
completely different feature — see src/lib/sgt-schema.ts).

The per-page routing/clustering (including cross-page table continuation)
lives in _ocr_common.py (ocr_layout_docx_pages()) so this script and the
persistent worker (ocr_worker.py) reuse the exact same evidence the .docx
path already uses — this file just renders those blocks to .xlsx instead.

Usage: python ocr_layout_to_xlsx.py <input.pdf|input.png|...> <output.xlsx> [--pages 0,2,5]
Prints OK on success. On failure, prints the traceback to stderr and exits non-zero.
"""
import sys

from _ocr_common import LazyModel, make_layout_detector, make_ocr, make_table_pipeline, ocr_layout_docx_pages


def write_xlsx(pages_blocks, output_path):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    multi_page = sum(1 for b in pages_blocks if b) > 1
    sheet_count = 0

    for blocks in pages_blocks:
        if not blocks:
            continue
        sheet_count += 1
        ws = wb.create_sheet(title=f"Page {sheet_count}" if multi_page else "Sheet1")
        row_cursor = 1

        for kind, payload in blocks:
            if kind == "table":
                rows = payload["grid"]
                spans = payload.get("spans", [])
                if not rows:
                    continue
                ncols = max(len(r) for r in rows)
                table_start_row = row_cursor

                for r_idx, row_cells in enumerate(rows):
                    for c_idx in range(ncols):
                        text = row_cells[c_idx] if c_idx < len(row_cells) else ""
                        ws.cell(row=table_start_row + r_idx, column=c_idx + 1, value=text)

                # Reproduce the source's real merged cells, same overlap-safe
                # approach as write_docx() in ocr_layout_to_docx.py — the
                # recognized spans can conflict on messy real-world tables,
                # so skip (rather than crash on) anything that overlaps a
                # cell already merged by an earlier span.
                merged_cells = set()
                for r0, c0, r1, c1 in sorted(spans, key=lambda s: (s[0], s[1])):
                    if r1 >= len(rows) or c1 >= ncols or (r0, c0) == (r1, c1):
                        continue
                    span_cells = {(r, c) for r in range(r0, r1 + 1) for c in range(c0, c1 + 1)}
                    if span_cells & merged_cells:
                        continue
                    try:
                        ws.merge_cells(
                            start_row=table_start_row + r0,
                            start_column=c0 + 1,
                            end_row=table_start_row + r1,
                            end_column=c1 + 1,
                        )
                        merged_cells |= span_cells
                    except Exception:  # noqa: BLE001 — a conflicting merge shouldn't lose the rest of the sheet
                        continue

                row_cursor = table_start_row + len(rows) + 1  # blank row separates table from what follows
            else:
                ws.cell(row=row_cursor, column=1, value=payload)
                row_cursor += 1

    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")

    wb.save(output_path)


def main() -> int:
    if len(sys.argv) < 3:
        print("usage: ocr_layout_to_xlsx.py <input> <output.xlsx> [--pages 0,2,5]", file=sys.stderr)
        return 2

    input_path, output_path = sys.argv[1], sys.argv[2]
    pages = None
    if "--pages" in sys.argv:
        idx = sys.argv.index("--pages")
        pages = [int(p) for p in sys.argv[idx + 1].split(",") if p.strip() != ""]

    ocr = make_ocr()
    layout_detector = make_layout_detector()
    table_pipeline = LazyModel(make_table_pipeline)

    try:
        pages_blocks = ocr_layout_docx_pages(ocr, layout_detector, table_pipeline, input_path, pages)
    except ValueError as e:
        print(str(e), file=sys.stderr)
        return 2

    if not any(pages_blocks):
        print("no text detected", file=sys.stderr)
        return 1

    write_xlsx(pages_blocks, output_path)
    print("OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
