"""End-to-end: scanned cargo-manifest PDF -> filled .xlsx matching the
source «Грузовой манифест» template.

Pipeline: PaddleOCR (ocr_document_pages) -> NVIDIA NIM text-LLM cleanup into
structured JSON (header + items) -> write values into the known Sheet1 cell
layout of templates/cargo-manifest-template.xlsx, inserting extra item rows
(copying row 14's style) when there are more items than the template has.

Usage:
  python fill_cargo_manifest.py <input.pdf> <output.xlsx> [--template path.xlsx]

Prints a JSON report {"warnings": [...], "itemCount": N} to stdout on success.
"""
import copy
import json
import re
import sqlite3
import sys
import urllib.error
import urllib.request

sys.path.insert(0, "scripts")
from _ocr_common import LazyModel, make_layout_detector, make_ocr, make_table_pipeline, ocr_document_pages  # noqa: E402

DEFAULT_TEMPLATE = r"E:\Файлы для распознавания\Грузовые манифесты\ГМ ЮКМ-ХЛМ-00031 - Ларга - 21.08.2024_модуль плавучести, контейнеры, отходы.xlsx"
NVIDIA_MODEL = "deepseek-ai/deepseek-v4-flash-0731"

SCHEMA = """{
  "header": {
    "portOfLoading": "string|null",
    "portOfDestination": "string|null",
    "vessel": "string|null",
    "shipper": "string|null",
    "consignee": "string|null",
    "departureDate": "string|null (as printed, e.g. 'Cp.21.08.2024')",
    "manifestNo": "string|null (format like ЮКМ-ХЛМ-00031 or ХЛМ-ЮКМ-00054)",
    "cargoPlacesCount": "string|null (plain integer)",
    "totalWeightKg": "string|null (plain number, dot decimal)",
    "onDeckWeightKg": "string|null",
    "bulkLiquidWeightKg": "string|null",
    "mtoRequest": "string|null"
  },
  "items": [
    {
      "no": "string (№ п/п, e.g. '1', '9.1')",
      "qty": "string (PLAIN NUMBER ONLY - if the OCR text shows a non-digit symbol here, infer the digit from context/quantity logic, never leave a stray letter/symbol)",
      "unit": "string",
      "lengthMm": "string|null (PLAIN NUMBER)",
      "widthMm": "string|null (PLAIN NUMBER)",
      "heightMm": "string|null (PLAIN NUMBER)",
      "placesCount": "string|null (PLAIN NUMBER ONLY - same rule as qty)",
      "description": "string",
      "weightKg": "string (PLAIN NUMBER)",
      "unitId": "string|null",
      "supplier": "string|null",
      "techDescription": "string|null",
      "hazardClass": "string|null",
      "documentNo": "string|null",
      "goodsNo": "string|null"
    }
  ]
}"""


def build_prompt(raw_text: str) -> str:
    return f"""You are cleaning up noisy OCR output of a Russian cargo manifest (Грузовой манифест) table and reconstructing it as clean structured JSON.

The OCR text below is imperfect. Known systematic OCR error patterns in this pipeline you must correct using context and domain knowledge:
- The first 1-2 characters of a cell are sometimes cut off (e.g. "орт погрузки" means "Порт погрузки", "удно" means "Судно", "рузоотправитель" means "Грузоотправитель").
- Cyrillic "Ю" is sometimes misread as Latin "IO" (e.g. "IOKM" should be "ЮКМ").
- Digits in numeric columns (Кол-во, Кол-во грузовых мест, Вес, Размер) are sometimes misread as similar-looking letters/symbols (e.g. "§" for "5", "ч" for "4", "а" for "4", ">" for "7"). These columns must ALWAYS end up as plain digits - if you see a non-digit in a numeric column, reason about what digit it visually resembles and/or what value makes the row's totals consistent, and use that digit. Never leave a letter or symbol in a numeric field.
- Company name OCR noise: keep proper nouns as printed if genuinely ambiguous, but resolve obvious partial/garbled forms using the rest of the document as context (the same company often appears on multiple rows).
- Rows can get column-shifted on multi-line cells - use column position and row logic (weight, then supplier, then techDescription, then hazard class, then document number "ГМ №...", then goods number) to realign values into the right field.
- The manifest number pattern is always like "XXX-YYY-NNNNN" using 3-4 letter port/route codes (e.g. ЮКМ, ХЛМ, НХД) - fix obviously garbled versions of this pattern using the pattern shape, not by inventing new codes.
- CRITICAL - do not duplicate/carry over values between rows: each row is independent. A row is very often missing dimensions, supplier, unitId or techDescription entirely (e.g. waste/liquid sub-items like "9.1" typically have qty+unit+description+weight+documentNo ONLY - no dimensions, no supplier, no unitId, no techDescription, no goodsNo). If a cell is not explicitly printed for THAT row, it is null - never copy a neighboring row's value into it just because the table looks visually continuous.
- techDescription and hazardClass are two DIFFERENT columns. Do not put the same text in both unless the source genuinely repeats it in both columns.

Rules:
- Return ONLY valid JSON matching this exact schema, no markdown fences, no commentary:
{SCHEMA}
- If a value is genuinely illegible/absent from the OCR text (not just noisy), use null. Do NOT invent values that aren't grounded in the OCR text.
- Transcribe cargo descriptions in full detail, don't truncate.
- items must include EVERY row found in the OCR text, including sub-rows like "9.1", across all pages (multi-page manifests continue the same table).
- Numeric weight/dimension fields: plain numbers, dot decimal separator, no thousands separators.

OCR TEXT (reading order, markdown table format, may span multiple pages):
---
{raw_text}
---

Return the corrected JSON now."""


def run_ocr(pdf_path: str) -> str:
    ocr = make_ocr()
    layout_detector = make_layout_detector()
    table_pipeline = LazyModel(make_table_pipeline)
    results = ocr_document_pages(ocr, layout_detector, table_pipeline, pdf_path, None)
    return "\n\n=== PAGE BREAK ===\n\n".join(p["text"] for p in results)


def run_llm(raw_text: str) -> dict:
    con = sqlite3.connect("db/custom.db")
    cur = con.cursor()
    cur.execute("select value from Settings where key='llm_api_key'")
    key = cur.fetchone()[0]

    body = {
        "model": NVIDIA_MODEL,
        "temperature": 0.0,
        "max_tokens": 8000,
        "messages": [{"role": "user", "content": build_prompt(raw_text)}],
    }
    req = urllib.request.Request(
        "https://integrate.api.nvidia.com/v1/chat/completions",
        data=json.dumps(body).encode(),
        headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=280) as r:
        data = json.load(r)
    content = data["choices"][0]["message"]["content"]
    match = re.search(r"```(?:json)?\s*([\s\S]*?)```", content)
    json_str = match.group(1) if match else content
    return json.loads(json_str.strip())


NUMERIC_FIELDS = {"qty", "placesCount", "weightKg", "lengthMm", "widthMm", "heightMm"}
NUMERIC_RE = re.compile(r"^-?\d+([.,]\d+)?$")


def validate(extracted: dict) -> list:
    warnings = []
    for idx, item in enumerate(extracted.get("items", [])):
        row_label = f"item #{item.get('no', idx)}"
        for field in NUMERIC_FIELDS:
            val = item.get(field)
            if val and not NUMERIC_RE.match(str(val).strip()):
                warnings.append(f"{row_label}: field '{field}' has non-numeric value {val!r} - needs manual check")
        if not item.get("description"):
            warnings.append(f"{row_label}: missing description")
    header = extracted.get("header", {})
    try:
        declared_total = float(str(header.get("totalWeightKg", "")).replace(",", "."))
        computed_total = sum(
            float(str(it["weightKg"]).replace(",", "."))
            for it in extracted.get("items", [])
            if it.get("weightKg") and NUMERIC_RE.match(str(it["weightKg"]).strip())
        )
        if abs(declared_total - computed_total) > 1.0:
            warnings.append(
                f"header totalWeightKg={declared_total} does not match sum of item weights={computed_total:.2f} - check item weight OCR"
            )
    except (TypeError, ValueError):
        pass
    return warnings


def fill_workbook(extracted: dict, template_path: str, output_path: str):
    from openpyxl import load_workbook

    wb = load_workbook(template_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]

    h = extracted.get("header", {})
    ws["C5"] = h.get("portOfLoading")
    ws["C6"] = h.get("portOfDestination")
    ws["C7"] = h.get("vessel")
    ws["C8"] = h.get("shipper")
    ws["C9"] = h.get("consignee")
    ws["C10"] = h.get("mtoRequest") or "-"
    ws["M5"] = h.get("departureDate")
    ws["M6"] = h.get("manifestNo")
    ws["M7"] = _num(h.get("cargoPlacesCount"))
    ws["M8"] = _num(h.get("totalWeightKg"))
    ws["M9"] = _num(h.get("onDeckWeightKg"))
    ws["M10"] = _num(h.get("bulkLiquidWeightKg"))

    items = extracted.get("items", [])
    FIRST_DATA_ROW = 14
    TEMPLATE_DATA_ROWS = 12  # rows 14..25 in the source template (11 numbered items incl. 9.1)
    needed_extra = max(0, len(items) - TEMPLATE_DATA_ROWS)

    if needed_extra:
        insert_at = FIRST_DATA_ROW + TEMPLATE_DATA_ROWS
        ws.insert_rows(insert_at, amount=needed_extra)
        style_row = FIRST_DATA_ROW
        for r_offset in range(needed_extra):
            target_row = insert_at + r_offset
            for col in range(1, 16):
                src = ws.cell(row=style_row, column=col)
                dst = ws.cell(row=target_row, column=col)
                dst._style = copy.copy(src._style)

    total_weight = 0.0
    total_places = 0.0
    for i, item in enumerate(items):
        row = FIRST_DATA_ROW + i
        ws.cell(row=row, column=1, value=item.get("no"))
        ws.cell(row=row, column=2, value=_num(item.get("qty"), keep_str_on_fail=True))
        ws.cell(row=row, column=3, value=item.get("unit"))
        ws.cell(row=row, column=4, value=_num(item.get("lengthMm")))
        ws.cell(row=row, column=5, value=_num(item.get("widthMm")))
        ws.cell(row=row, column=6, value=_num(item.get("heightMm")))
        ws.cell(row=row, column=7, value=_num(item.get("placesCount"), keep_str_on_fail=True))
        ws.cell(row=row, column=8, value=item.get("description"))
        w = _num(item.get("weightKg"), keep_str_on_fail=True)
        ws.cell(row=row, column=9, value=w)
        ws.cell(row=row, column=10, value=item.get("unitId"))
        ws.cell(row=row, column=11, value=item.get("supplier"))
        ws.cell(row=row, column=12, value=item.get("techDescription"))
        ws.cell(row=row, column=13, value=item.get("hazardClass"))
        ws.cell(row=row, column=14, value=item.get("documentNo"))
        ws.cell(row=row, column=15, value=item.get("goodsNo"))
        if isinstance(w, (int, float)):
            total_weight += w
        p = _num(item.get("placesCount"))
        if isinstance(p, (int, float)):
            total_places += p

    # Clear any leftover template rows beyond what was actually extracted
    # (e.g. template has capacity for 12 items but only 11 were found) so
    # stale sample data from the source template doesn't survive into the
    # output.
    if len(items) < TEMPLATE_DATA_ROWS:
        for r in range(FIRST_DATA_ROW + len(items), FIRST_DATA_ROW + TEMPLATE_DATA_ROWS):
            for col in range(1, 16):
                ws.cell(row=r, column=col, value=None)

    totals_row = FIRST_DATA_ROW + len(items) + (1 if len(items) < TEMPLATE_DATA_ROWS else 0)
    # best-effort: only touch the totals row if it looks like the original template's totals row
    # (column H says "Всего вес груза" or similar) - otherwise leave alone.
    label_cell = ws.cell(row=totals_row, column=8).value
    if label_cell and ("вес" in str(label_cell).lower() or "всего" in str(label_cell).lower()):
        ws.cell(row=totals_row, column=7, value=total_places or None)
        ws.cell(row=totals_row, column=9, value=round(total_weight, 2) or None)

    wb.save(output_path)


def _num(val, keep_str_on_fail=False):
    if val is None:
        return None
    s = str(val).strip().replace(",", ".").replace(" ", "")
    if s in ("", "-"):
        return None
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return val if keep_str_on_fail else None


def main():
    if len(sys.argv) < 3:
        print("usage: fill_cargo_manifest.py <input.pdf> <output.xlsx> [--template path]", file=sys.stderr)
        return 2
    pdf_path, out_path = sys.argv[1], sys.argv[2]
    template_path = DEFAULT_TEMPLATE
    if "--template" in sys.argv:
        template_path = sys.argv[sys.argv.index("--template") + 1]

    raw_text = run_ocr(pdf_path)
    extracted = run_llm(raw_text)
    warnings = validate(extracted)
    fill_workbook(extracted, template_path, out_path)

    print(json.dumps({"warnings": warnings, "itemCount": len(extracted.get("items", []))}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
